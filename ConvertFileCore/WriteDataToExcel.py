import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, colors
import time

from ConvertFileCore.GetXmlTestcaseList import GetXmlTestcaseList


class WriteDataToExcel:
    def WriteData(self, CaseDatas):
        wb = Workbook()
        # 激活 worksheet
        ws = wb.active

        font = Font(name=u'宋体', size=10, color=colors.BLACK, bold=False)
        alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
        alignment_left = Alignment(horizontal='left', vertical='center', wrapText=True)
        fill_green = PatternFill(fill_type='solid', fgColor="228B22")

        for CaseData in CaseDatas:
            # print(CaseData)
            if 'testcase' in CaseData and len(CaseData['testcase']) > 0:
                testCases = CaseData['testcase']
                #创建sheet页，名字为suite名字
                ws = wb.create_sheet(CaseData['suite_name'])
                # 填充内容
                ws['A1'] = '标题'
                ws['B1'] = '预置条件'
                ws['C1'] = '步骤'
                ws['D1'] = '预期结果'
                #设置字体
                ws['A1'].font = font
                ws['B1'].font = font
                ws['C1'].font = font
                ws['D1'].font = font
                #填充背景颜色
                ws['A1'].fill = fill_green
                ws['B1'].fill = fill_green
                ws['C1'].fill = fill_green
                ws['D1'].fill = fill_green
                #文字居中
                ws['A1'].alignment = alignment_center
                ws['B1'].alignment = alignment_center
                ws['C1'].alignment = alignment_center
                ws['D1'].alignment = alignment_center
                # 调整列宽
                ws.column_dimensions['A'].width = 30.0
                ws.column_dimensions['B'].width = 20.0
                ws.column_dimensions['C'].width = 60.0
                ws.column_dimensions['D'].width = 80.0

                row = 2
                for testCase in testCases:
                    # 调整行高
                    # ws.row_dimensions[row].height = 40.0
                    #将case写入excel
                    if 'case_name' in testCase and len(testCase['case_name']) > 0:
                        ws['A' + str(row)] = testCase['case_name']
                        ws['A' + str(row)].font = font
                        ws['A' + str(row)].alignment = alignment_left
                    if 'case_summary' in testCase and len(testCase['case_summary']) > 0:
                        ws['B' + str(row)] = testCase['case_summary']
                        ws['B' + str(row)].font = font
                        ws['B' + str(row)].alignment = alignment_left
                    if 'case_steps' in testCase and len(testCase['case_steps']) > 0:
                        ws['C' + str(row)] = testCase['case_steps']
                        ws['C' + str(row)].font = font
                        ws['C' + str(row)].alignment = alignment_left
                    if 'case_expectedresults' in testCase and len(testCase['case_expectedresults']) > 0:
                        ws['D' + str(row)] = testCase['case_expectedresults']
                        ws['D' + str(row)].font = font
                        ws['D' + str(row)].alignment = alignment_left
                    row += 1

        # 可以附加行，从第一列开始附加
        # ws.append([1, 2, 3])
        # 保存文件
        wb.save("TestCase"+str(int(time.time()))+".xlsx")


if __name__ == '__main__':
    caselist = GetXmlTestcaseList().ReadCaseFromXml('C:\\Users\\Xone\\Desktop\\all_testsuites.xml')
    WriteDataToExcel().WriteData(caselist)
